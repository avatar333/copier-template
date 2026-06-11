from __future__ import annotations

from io import BytesIO
from datetime import date
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from app.extensions import db
from app.models import AppUser, AssignmentRun, HostAssignment
from app.services.exporter import (
    ExportError,
    ExportResult,
    ExportedWorkbook,
    ExportZipResult,
    export_assignment_run_to_excel,
    export_assignment_run_to_zip,
    find_hostname_column,
    sanitize_filename_part,
)


def test_sanitize_filename_part_strips_unsafe_characters():
    assert sanitize_filename_part("  CRQ 12/34 (Final)! ") == "crq-12-34-final"


def test_exporter_finds_hostname_column_in_template():
    sheet_target = find_hostname_column(Path("TEMPLATE - Patching run tracking and logging.xltx"))

    assert sheet_target.path.endswith("sheet1.xml")
    assert sheet_target.hostname_column == "B"
    assert "Dev" in sheet_target.title or "Stage" in sheet_target.title or "Production" in sheet_target.title


def test_export_assignment_run_to_excel_writes_sorted_hosts_and_skips_empty_users(app, tmp_path):
    with app.app_context():
        admin = AppUser.query.filter_by(login_name="admin").first()
        extra_user = AppUser(login_name="alice", first_name="Alice", last_name="Able", is_active=True)
        extra_user.set_password("password123")
        db.session.add(extra_user)
        db.session.flush()

        run = AssignmentRun(
            created_by_user_id=admin.id,
            pool_name="production",
            pool_collection_name="Production",
            host_count=3,
            user_count=3,
        )
        db.session.add(run)
        db.session.flush()
        db.session.add_all(
            [
                HostAssignment(
                    assignment_run_id=run.id,
                    user_id=admin.id,
                    fqdn="zulu02.example.com",
                    source_type="random",
                    source_name=None,
                ),
                HostAssignment(
                    assignment_run_id=run.id,
                    user_id=admin.id,
                    fqdn="alpha01.example.com",
                    source_type="pet",
                    source_name=None,
                ),
                HostAssignment(
                    assignment_run_id=run.id,
                    user_id=extra_user.id,
                    fqdn="bravo03.example.com",
                    source_type="prefix_sequence",
                    source_name="example-group",
                ),
            ]
        )
        db.session.commit()

        result = export_assignment_run_to_excel(run.id, "CRQ 12/34", tmp_path)

    assert isinstance(result, ExportResult)
    assert result.sheet_target.hostname_column == "B"
    assert len(result.files) == 2
    assert result.skipped_users == []

    exported_names = {item.filename for item in result.files}
    today = date.today().isoformat()
    assert f"admin_crq-12-34_{today}.xlsx" in exported_names
    assert f"alice_crq-12-34_{today}.xlsx" in exported_names

    admin_export = next(item for item in result.files if item.user_display_name == "admin")
    alice_export = next(item for item in result.files if item.user_display_name == "alice")
    assert admin_export.file_path.exists()
    assert alice_export.file_path.exists()

    assert _read_hostnames(admin_export.file_path, result.sheet_target.path) == [
        "alpha01.example.com",
        "zulu02.example.com",
    ]
    assert _read_hostnames(alice_export.file_path, result.sheet_target.path) == [
        "bravo03.example.com",
    ]
    workbook = load_workbook(admin_export.file_path)
    worksheet = workbook[workbook.sheetnames[0]]
    assert worksheet["B4"].value == "alpha01.example.com"
    assert worksheet["B4"].font.bold is False
    _assert_export_validations(worksheet)
    assert _is_normal_xlsx_workbook(admin_export.file_path)


def test_export_assignment_run_overwrites_existing_files_safely(app, tmp_path):
    with app.app_context():
        admin = AppUser.query.filter_by(login_name="admin").first()
        run = AssignmentRun(
            created_by_user_id=admin.id,
            pool_name="production",
            pool_collection_name="Production",
            host_count=1,
            user_count=2,
        )
        db.session.add(run)
        db.session.flush()
        db.session.add(
            HostAssignment(
                assignment_run_id=run.id,
                user_id=admin.id,
                fqdn="overwrite01.example.com",
                source_type="random",
                source_name=None,
            )
        )
        db.session.commit()

        expected_path = tmp_path / f"admin_crq-88_{date.today().isoformat()}.xlsx"
        expected_path.write_bytes(b"old-data")

        result = export_assignment_run_to_excel(run.id, "CRQ 88", tmp_path)

    assert expected_path.exists()
    assert expected_path.stat().st_size > len(b"old-data")
    assert _read_hostnames(expected_path, result.sheet_target.path) == ["overwrite01.example.com"]


def test_export_assignment_run_handles_missing_user_fallback(app, tmp_path):
    with app.app_context():
        admin = AppUser.query.filter_by(login_name="admin").first()
        run = AssignmentRun(
            created_by_user_id=admin.id,
            pool_name="production",
            pool_collection_name="Production",
            host_count=1,
            user_count=1,
        )
        db.session.add(run)
        db.session.flush()
        db.session.add(
            HostAssignment(
                assignment_run_id=run.id,
                user_id=None,
                fqdn="orphan01.example.com",
                source_type="random",
                source_name=None,
            )
        )
        db.session.commit()

        result = export_assignment_run_to_excel(run.id, "CRQ 99", tmp_path)

    assert len(result.files) == 1
    assert result.files[0].filename.startswith("assignment-user-unknown_crq-99_")
    assert _read_hostnames(result.files[0].file_path, result.sheet_target.path) == ["orphan01.example.com"]


def test_export_assignment_run_to_zip_creates_downloadable_zip_with_xlsx_files(app, tmp_path):
    with app.app_context():
        admin = AppUser.query.filter_by(login_name="admin").first()
        alice = AppUser(login_name="alice", first_name="Alice", last_name="Able", is_active=True)
        alice.set_password("password123")
        db.session.add(alice)
        db.session.flush()

        run = AssignmentRun(
            created_by_user_id=admin.id,
            pool_name="production",
            pool_collection_name="Production",
            host_count=2,
            user_count=2,
        )
        db.session.add(run)
        db.session.flush()
        db.session.add_all(
            [
                HostAssignment(
                    assignment_run_id=run.id,
                    user_id=admin.id,
                    fqdn="zulu02.example.com",
                    source_type="random",
                    source_name=None,
                ),
                HostAssignment(
                    assignment_run_id=run.id,
                    user_id=admin.id,
                    fqdn="alpha01.example.com",
                    source_type="pet",
                    source_name=None,
                ),
                HostAssignment(
                    assignment_run_id=run.id,
                    user_id=alice.id,
                    fqdn="bravo03.example.com",
                    source_type="prefix_sequence",
                    source_name="example-group",
                ),
            ]
        )
        db.session.commit()

        result = export_assignment_run_to_zip(run.id, "CRQ 12/34", tmp_path)

    assert isinstance(result, ExportZipResult)
    assert result.zip_path.exists()
    assert result.zip_filename == f"assignment_run_{run.id}_crq-12-34_{date.today().isoformat()}.zip"

    with ZipFile(result.zip_path) as archive:
        names = archive.namelist()
        assert all(name.endswith(".xlsx") for name in names)
        assert len(names) == 2
        assert f"admin_crq-12-34_{date.today().isoformat()}.xlsx" in names
        assert f"alice_crq-12-34_{date.today().isoformat()}.xlsx" in names

        admin_name = next(name for name in names if name.startswith("admin_"))
        with archive.open(admin_name) as workbook_bytes:
            workbook = load_workbook(BytesIO(workbook_bytes.read()))
            worksheet = workbook[workbook.sheetnames[0]]
            assert worksheet["B4"].value == "alpha01.example.com"
            assert worksheet["B4"].font.bold is False
            _assert_export_validations(worksheet)


def test_export_assignment_run_fails_cleanly_for_missing_template(tmp_path):
    with pytest.raises(ExportError, match="Export template not found"):
        export_assignment_run_to_excel(1, "CRQ 100", tmp_path, template_path=tmp_path / "missing.xltx")


def test_export_assignment_run_fails_cleanly_for_missing_values_sheet(app, tmp_path):
    bad_template = tmp_path / "missing_values.xlsx"
    _copy_template_without_values(Path("TEMPLATE - Patching run tracking and logging.xltx"), bad_template)

    with app.app_context():
        admin = AppUser.query.filter_by(login_name="admin").first()
        run = AssignmentRun(
            created_by_user_id=admin.id,
            pool_name="production",
            pool_collection_name="Production",
            host_count=1,
            user_count=1,
        )
        db.session.add(run)
        db.session.flush()
        db.session.add(
            HostAssignment(
                assignment_run_id=run.id,
                user_id=admin.id,
                fqdn="orphan01.example.com",
                source_type="random",
                source_name=None,
            )
        )
        db.session.commit()

        with pytest.raises(ExportError, match="Values"):
            export_assignment_run_to_excel(run.id, "CRQ 102", tmp_path, template_path=bad_template)


def test_export_assignment_run_fails_cleanly_for_missing_target_sheet(app, tmp_path):
    bad_template = tmp_path / "missing_target.xlsx"
    _copy_template_without_target_sheet(Path("TEMPLATE - Patching run tracking and logging.xltx"), bad_template)

    with app.app_context():
        admin = AppUser.query.filter_by(login_name="admin").first()
        run = AssignmentRun(
            created_by_user_id=admin.id,
            pool_name="production",
            pool_collection_name="Production",
            host_count=1,
            user_count=1,
        )
        db.session.add(run)
        db.session.flush()
        db.session.add(
            HostAssignment(
                assignment_run_id=run.id,
                user_id=admin.id,
                fqdn="orphan01.example.com",
                source_type="random",
                source_name=None,
            )
        )
        db.session.commit()

        with pytest.raises(ExportError, match="target worksheet"):
            export_assignment_run_to_excel(run.id, "CRQ 103", tmp_path, template_path=bad_template)


def test_export_assignment_run_to_zip_fails_when_no_files_are_generated(app, tmp_path):
    with app.app_context():
        admin = AppUser.query.filter_by(login_name="admin").first()
        run = AssignmentRun(
            created_by_user_id=admin.id,
            pool_name="production",
            pool_collection_name="Production",
            host_count=0,
            user_count=1,
        )
        db.session.add(run)
        db.session.commit()

        with pytest.raises(ExportError, match="No assigned hosts were available to export"):
            export_assignment_run_to_zip(run.id, "CRQ 104", tmp_path)


def _read_hostnames(workbook_path: Path, sheet_path: str) -> list[str]:
    workbook = load_workbook(workbook_path)
    worksheet = workbook[sheet_path] if sheet_path in workbook.sheetnames else workbook.worksheets[0]
    values = []
    for cell in worksheet["B"][3:]:
        if cell.value:
            values.append(cell.value)
    return values


def _is_normal_xlsx_workbook(workbook_path: Path) -> bool:
    with ZipFile(workbook_path) as workbook_zip:
        content_types = workbook_zip.read("[Content_Types].xml").decode()
    return (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
        in content_types
        and "application/vnd.openxmlformats-officedocument.spreadsheetml.template.main+xml"
        not in content_types
    )


def _copy_template_without_target_sheet(source_template: Path, destination: Path) -> None:
    workbook = load_workbook(source_template)
    if workbook.sheetnames:
        workbook.remove(workbook[workbook.sheetnames[0]])
    workbook.save(destination)


def _copy_template_without_values(source_template: Path, destination: Path) -> None:
    workbook = load_workbook(source_template)
    if "Values" in workbook.sheetnames:
        workbook.remove(workbook["Values"])
    workbook.save(destination)


def _assert_export_validations(worksheet) -> None:
    validations = {str(dv.sqref): dv for dv in worksheet.data_validations.dataValidation}
    assert validations["D4:D100"].formula1 == "='Values'!$A$3:$A$10"
    assert validations["E4:E100"].formula1 == "='Values'!$B$3:$B$10"
    assert validations["F4:F100"].formula1 == "='Values'!$C$3:$C$10"
    assert validations["G4:G100"].formula1 == "='Values'!$D$3:$D$10"
    assert all(not dv.showInputMessage for dv in validations.values())
