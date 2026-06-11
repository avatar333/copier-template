from __future__ import annotations

import os
import re
import tempfile
import warnings
from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from ..extensions import db
from ..models import AssignmentRun, AppUser

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

TEMPLATE_FILENAME = "TEMPLATE - Patching run tracking and logging.xltx"
HOSTNAME_START_ROW = 4
HOSTNAME_COLUMN = "B"
DATA_VALIDATION_RANGES = {
    "D": "'Values'!$A$3:$A$10",
    "E": "'Values'!$B$3:$B$10",
    "F": "'Values'!$C$3:$C$10",
    "G": "'Values'!$D$3:$D$10",
}


class ExportError(RuntimeError):
    """Raised when an assignment run cannot be exported."""


@dataclass(frozen=True)
class SheetTarget:
    title: str
    path: str
    hostname_column: str


@dataclass(frozen=True)
class ExportedWorkbook:
    file_path: Path
    filename: str
    user_display_name: str
    host_count: int


@dataclass(frozen=True)
class ExportResult:
    files: list[ExportedWorkbook]
    skipped_users: list[str]
    sheet_target: SheetTarget


@dataclass(frozen=True)
class ExportZipResult:
    zip_path: Path
    zip_filename: str
    files: list[ExportedWorkbook]
    skipped_users: list[str]
    sheet_target: SheetTarget


def sanitize_filename_part(value: str | None) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    text = re.sub(r"\s+", "-", text)
    text = re.sub(r"[^a-z0-9._-]+", "-", text)
    text = re.sub(r"-{2,}", "-", text)
    text = re.sub(r"\.{2,}", ".", text)
    text = re.sub(r"_{2,}", "_", text)
    return text.strip("._-")


def find_hostname_column(workbook_path: Path | str) -> SheetTarget:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise ExportError(f"Export template not found: {workbook_path.name}")

    workbook = _load_template_workbook(workbook_path)
    worksheet = _select_target_worksheet(workbook, None)
    if worksheet is None:
        raise ExportError("The export template does not contain the expected target worksheet.")
    return SheetTarget(
        title=worksheet.title,
        path=_sheet_path_for_title(workbook_path, worksheet.title),
        hostname_column=HOSTNAME_COLUMN,
    )


def export_assignment_run_to_excel(
    run_id: int,
    change_request_number: str,
    output_dir: Path | str,
    template_path: Path | str | None = None,
) -> ExportResult:
    normalized_change_request = sanitize_filename_part(change_request_number)
    if not normalized_change_request:
        raise ExportError("Change Request Number is required.")

    template_path = Path(template_path or _default_template_path())
    if not template_path.exists():
        raise ExportError(f"Export template not found: {template_path.name}")

    run = db.session.get(AssignmentRun, run_id)
    if run is None:
        raise ExportError("The requested assignment run could not be found.")

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = _load_template_workbook(template_path)
    worksheet = _select_target_worksheet(workbook, pool_name=run.pool_name)
    if worksheet is None:
        raise ExportError(
            "The export template does not contain the expected target worksheet for this assignment pool."
        )
    sheet_target = SheetTarget(
        title=worksheet.title,
        path=_sheet_path_for_title(template_path, worksheet.title),
        hostname_column=HOSTNAME_COLUMN,
    )

    assignments_by_user_id: dict[int | None, list[str]] = {}
    for assignment in sorted(run.assignments, key=lambda item: (item.user_id or 0, item.fqdn)):
        assignments_by_user_id.setdefault(assignment.user_id, []).append(assignment.fqdn)

    file_date = date.today().isoformat()
    exported_files: list[ExportedWorkbook] = []
    skipped_users: list[str] = []
    temp_paths: list[Path] = []
    try:
        for user_id, hostnames in sorted(assignments_by_user_id.items(), key=_assignment_sort_key):
            sorted_hosts = sorted({host.strip().lower() for host in hostnames if host and host.strip()})
            if not sorted_hosts:
                skipped_users.append(_fallback_user_label(user_id))
                continue
            user_label = _resolve_user_label(user_id)
            filename = _build_filename(user_label, normalized_change_request, file_date)
            final_path = output_dir / filename
            with tempfile.NamedTemporaryFile(
                mode="wb",
                suffix=".tmp",
                prefix=f"{filename}.",
                dir=output_dir,
                delete=False,
            ) as temp_file:
                temp_path = Path(temp_file.name)
                temp_paths.append(temp_path)
            _write_workbook(template_path, temp_path, sheet_target, sorted_hosts)
            os.replace(temp_path, final_path)
            temp_paths.remove(temp_path)
            exported_files.append(
                ExportedWorkbook(
                    file_path=final_path,
                    filename=filename,
                    user_display_name=user_label,
                    host_count=len(sorted_hosts),
                )
            )
        return ExportResult(files=exported_files, skipped_users=skipped_users, sheet_target=sheet_target)
    except Exception:
        for temp_path in temp_paths:
            try:
                temp_path.unlink(missing_ok=True)
            except OSError:
                pass
        raise


def export_assignment_run_to_zip(
    run_id: int,
    change_request_number: str,
    output_dir: Path | str,
    template_path: Path | str | None = None,
) -> ExportZipResult:
    normalized_change_request = sanitize_filename_part(change_request_number)
    if not normalized_change_request:
        raise ExportError("Change Request Number is required.")

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    export_result = export_assignment_run_to_excel(
        run_id,
        change_request_number,
        output_dir,
        template_path=template_path,
    )
    if not export_result.files:
        raise ExportError("No assigned hosts were available to export for this run.")

    zip_filename = _build_zip_filename(run_id, normalized_change_request, date.today().isoformat())
    zip_path = output_dir / zip_filename
    temp_zip_path = zip_path.with_suffix(".zip.tmp")
    try:
        with ZipFile(temp_zip_path, "w", ZIP_DEFLATED) as zip_file:
            for exported_file in export_result.files:
                zip_file.write(exported_file.file_path, arcname=exported_file.filename)
        os.replace(temp_zip_path, zip_path)
    except Exception:
        try:
            temp_zip_path.unlink(missing_ok=True)
        except OSError:
            pass
        raise

    return ExportZipResult(
        zip_path=zip_path,
        zip_filename=zip_filename,
        files=export_result.files,
        skipped_users=export_result.skipped_users,
        sheet_target=export_result.sheet_target,
    )


def _default_template_path() -> Path:
    return Path(__file__).resolve().parents[2] / TEMPLATE_FILENAME


def _build_filename(user_label: str, change_request: str, file_date: str) -> str:
    user_part = sanitize_filename_part(user_label) or "assignment-user"
    return f"{user_part}_{change_request}_{file_date}.xlsx"


def _build_zip_filename(run_id: int, change_request: str, file_date: str) -> str:
    return f"assignment_run_{run_id}_{change_request}_{file_date}.zip"


def _assignment_sort_key(entry: tuple[int | None, list[str]]) -> tuple[int, str]:
    user_id, _ = entry
    if user_id is None:
        return (10**9, "assignment-user-unknown")
    return (int(user_id), _fallback_user_label(user_id))


def _resolve_user_label(user_id: int | None) -> str:
    if user_id is None:
        return _fallback_user_label(None)
    user = db.session.get(AppUser, user_id)
    if user is None:
        return _fallback_user_label(user_id)
    login_name = sanitize_filename_part(user.login_name)
    if login_name:
        return login_name
    fallback_name = sanitize_filename_part(f"{user.first_name}.{user.last_name}")
    if fallback_name:
        return fallback_name
    return _fallback_user_label(user_id)


def _fallback_user_label(user_id: int | None) -> str:
    if user_id is None:
        return "assignment-user-unknown"
    return f"assignment-user-{user_id}"


def _write_workbook(
    template_path: Path,
    output_path: Path,
    sheet_target: SheetTarget,
    hosts: list[str],
) -> None:
    workbook = _load_template_workbook(template_path)
    workbook.template = False
    worksheet = workbook[sheet_target.title]

    _clear_hostname_values(worksheet, HOSTNAME_COLUMN, HOSTNAME_START_ROW)
    _apply_data_validations(worksheet)

    template_cell = worksheet[f"{HOSTNAME_COLUMN}{HOSTNAME_START_ROW}"]
    for offset, hostname in enumerate(hosts):
        row_number = HOSTNAME_START_ROW + offset
        target_cell = worksheet[f"{HOSTNAME_COLUMN}{row_number}"]
        target_cell.value = hostname
        _copy_basic_style(template_cell, target_cell)
        font = copy(target_cell.font)
        font.bold = False
        target_cell.font = font

    workbook.save(output_path)


def _load_template_workbook(template_path: Path):
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Data Validation extension is not supported.*")
        return load_workbook(template_path)


def _clear_hostname_values(worksheet, hostname_column: str, start_row: int) -> None:
    max_row = max(worksheet.max_row, 100)
    for row_number in range(start_row, max_row + 1):
        worksheet[f"{hostname_column}{row_number}"].value = None


def _copy_basic_style(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)


def _apply_data_validations(worksheet) -> None:
    workbook = worksheet.parent
    if "Values" not in workbook.sheetnames:
        raise ExportError("The export template is missing the Values worksheet.")
    _remove_existing_data_validations(worksheet)
    for column, formula in DATA_VALIDATION_RANGES.items():
        validation = DataValidation(
            type="list",
            formula1=f"={formula}",
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=False,
        )
        validation.errorTitle = "Invalid value"
        validation.error = "Select a value from the list."
        worksheet.add_data_validation(validation)
        validation.add(f"{column}4:{column}100")


def _remove_existing_data_validations(worksheet) -> None:
    existing = list(getattr(worksheet.data_validations, "dataValidation", []))
    if not existing:
        return
    worksheet.data_validations.dataValidation = [
        validation
        for validation in existing
        if not _validation_matches_export_ranges(validation)
    ]


def _validation_matches_export_ranges(validation) -> bool:
    ranges = str(getattr(validation, "sqref", ""))
    return any(f"{column}4:{column}100" in ranges for column in DATA_VALIDATION_RANGES)


def _select_target_worksheet(workbook, pool_name: str | None):
    normalized_pool = str(pool_name or "").strip().lower()
    if normalized_pool == "production":
        preferred_titles = ("Production-",)
    elif normalized_pool == "non_production":
        preferred_titles = ("<Dev|Stage|Production>-", "Stage-", "Dev-")
    else:
        preferred_titles = ("<Dev|Stage|Production>-",)

    for title in preferred_titles:
        if title in workbook.sheetnames:
            return workbook[title]

    for worksheet in workbook.worksheets:
        if worksheet.title.endswith("-"):
            return worksheet
    return None


def _sheet_path_for_title(workbook_path: Path, sheet_title: str) -> str:
    with ZipFile(workbook_path) as workbook_zip:
        workbook_root = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
        rels_root = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
        rels = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_root.findall(f"{{{PKG_REL_NS}}}Relationship")
        }
        for sheet in workbook_root.findall(f".//{{{MAIN_NS}}}sheet"):
            if sheet.attrib.get("name") != sheet_title:
                continue
            rel_id = sheet.attrib.get(f"{{{OFFICE_REL_NS}}}id")
            target = rels.get(rel_id or "")
            if not target:
                break
            return target if target.startswith("xl/") else f"xl/{target}"
    return sheet_title
